
# Hovedscript som koordinerer hele screeningen.
# Henter data via api.py, beregner scorer via scoring.py,
# og lagrer det sorterte resultatet til en Excel-fil i output-mappen.

import os
import logging
import pandas as pd
import numpy as np
from config import FAKTORER, MARKEDER, OUTPUT_MAPPE, OUTPUT_FILNAVN, TEST_ANTALL
from api import hent_instrumenter, hent_nøkkeltall, hent_kvartalstall, hent_kurser
from scoring import beregn_samlet_score
from rich.console import Console
from rich.table import Table
from logg import legg_til_logg

# Setter opp logging slik at vi får tydelige meldinger underveis
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)
console = Console()


def beregn_snitt_daglig_omsetning(kurser: pd.DataFrame, dager: int = 60) -> float:
    """
    Beregner gjennomsnittlig daglig omsetning i lokal valuta over siste N handelsdager.
    Omsetning = volum × sluttkurs.
    """
    if kurser.empty or len(kurser) < dager:
        return None
    kurser = kurser.sort_values("d", ascending=False)
    siste = kurser.head(dager)
    return (siste["v"] * siste["c"]).mean()


def beregn_kursmomentum(kurser: pd.DataFrame) -> float:
    """
    Beregner 12-1 måneder kursmomemtum.
    Formel: (Kurs for 1 måned siden / Kurs for 12 måneder siden) - 1
    Ekskluderer siste måned for å unngå reverseringseffekt.
    """
    if kurser.empty or len(kurser) < 252:
        return np.nan

    # Sorter på dato – nyeste først
    kurser = kurser.sort_values("d", ascending=False)
    kurs_1m = kurser.iloc[21]["c"]   # Kurs for ca. 1 måned siden
    kurs_12m = kurser.iloc[252]["c"] # Kurs for ca. 12 måneder siden

    return (kurs_1m / kurs_12m) - 1


def beregn_inntjeningsstreak(kvartal: pd.DataFrame) -> int:
    """
    Teller antall kvartaler på rad med positiv YoY-vekst.
    Gjøres for omsetning, EBITDA, EBIT og EPS.
    Streaken brytes så snart ett kvartal viser negativ vekst.
    """
    if kvartal.empty or len(kvartal) < 5:
        return 0

    # Sorter på år og kvartal – nyeste først
    kvartal = kvartal.sort_values(["year", "period"], ascending=False).reset_index(drop=True)

    metrikker = {
        "revenues": "omsetning",
        "ebitda": "ebitda",
        "ebit": "ebit",
        "eps": "eps",
    }

    min_streak = None

    for kolonne, navn in metrikker.items():
        if kolonne not in kvartal.columns:
            continue

        streak = 0
        for i in range(len(kvartal) - 4):
            nåværende = kvartal.iloc[i][kolonne]
            samme_kvartal_ifjor = kvartal.iloc[i + 4][kolonne]

            # Ekskluder EPS hvis negativ
            if kolonne == "eps" and (nåværende < 0 or samme_kvartal_ifjor < 0):
                break

            if nåværende > samme_kvartal_ifjor:
                streak += 1
            else:
                break

        if min_streak is None or streak < min_streak:
            min_streak = streak

    return min_streak if min_streak is not None else 0


def beregn_vekst_metrikker(kvartal: pd.DataFrame) -> dict:
    """
    Beregner YoY-vekst og akselerasjon for siste kvartal.
    Sammenligner siste kvartal med samme kvartal året før.
    Akselerasjon = er veksten bedre enn forrige kvartal?
    """
    if kvartal.empty or len(kvartal) < 5:
        return {}

    kvartal = kvartal.sort_values(["year", "period"], ascending=False).reset_index(drop=True)

    metrikker = {
        "revenues": "vekst_omsetning",
        "ebitda": "vekst_ebitda",
        "ebit": "vekst_ebit",
        "eps": "vekst_eps",
    }

    resultat = {}
    akselerasjon_verdier = []

    for kolonne, navn in metrikker.items():
        if kolonne not in kvartal.columns or len(kvartal) < 8:
            continue

        # Siste kvartal vs samme kvartal ifjor
        q0 = kvartal.iloc[0][kolonne]
        q0_ifjor = kvartal.iloc[4][kolonne]

        # Forrige kvartal vs samme kvartal ifjor
        q1 = kvartal.iloc[1][kolonne]
        q1_ifjor = kvartal.iloc[5][kolonne]

        # Ekskluder EPS hvis negativ
        if kolonne == "eps" and (q0 < 0 or q0_ifjor < 0):
            continue

        if q0_ifjor != 0:
            vekst_q0 = (q0 - q0_ifjor) / abs(q0_ifjor)
            resultat[navn] = vekst_q0

        if q1_ifjor != 0:
            vekst_q1 = (q1 - q1_ifjor) / abs(q1_ifjor)
            # Akselerasjon: 1 hvis veksten er bedre enn forrige kvartal, 0 hvis ikke
            if navn in resultat:
                akselerasjon_verdier.append(1 if resultat[navn] > vekst_q1 else 0)

    # Gjennomsnittlig akselerasjon på tvers av metrikker
    resultat["akselerasjon"] = np.mean(akselerasjon_verdier) if akselerasjon_verdier else np.nan

    return resultat


def hent_nøkkeltall_for_selskap(instrument_id: int) -> dict:
    """
    Henter og strukturerer nøkkeltall for ett selskap.
    Returnerer en dict med alle relevante metrikker.
    """
    return hent_nøkkeltall(instrument_id)


def kjør_screening():
    """
    Hovedfunksjon som kjører hele screeningen.
    1. Henter alle selskaper
    2. Henter data per selskap
    3. Beregner scores
    4. Lagrer output til CSV
    """
    logger.info("=== Starter faktorscreening ===")

    # Lag output-mappe hvis den ikke finnes
    os.makedirs(OUTPUT_MAPPE, exist_ok=True)

    # Steg 1: Hent alle selskaper
    instrumenter = hent_instrumenter()
    if instrumenter.empty:
        logger.error("Ingen selskaper hentet – avslutter")
        return

    # Steg 2: Hent data per selskap
    if TEST_ANTALL is not None:
        instrumenter = instrumenter.groupby("land").head(TEST_ANTALL).reset_index(drop=True)
        logger.info(f"TEST_ANTALL={TEST_ANTALL} per marked – kjører med begrenset utvalg")

    alle_data = []
    totalt = len(instrumenter)

    from rich.progress import Progress, BarColumn, TextColumn, TimeRemainingColumn
    with Progress(
        TextColumn("[bold blue]Henter data"),
        BarColumn(),
        TextColumn("{task.completed}/{task.total}"),
        TimeRemainingColumn(),
    ) as progress:
        task = progress.add_task("", total=totalt)

        for i, (_, selskap) in enumerate(instrumenter.iterrows()):
            instrument_id = selskap["insId"]
            progress.update(task, advance=1, description=f"[bold blue]{selskap.get('name', instrument_id)[:30]}")

            rad = {
                "navn":      selskap.get("name"),
                "ticker":    selskap.get("ticker"),
                "sektor":    selskap.get("sektor"),
                "land":      selskap.get("land"),
            }

            # Hent nøkkeltall
            nøkkeltall = hent_nøkkeltall_for_selskap(instrument_id)
            rad.update(nøkkeltall)

            # Hent kvartalstall og beregn vekst og streak
            kvartal = hent_kvartalstall(instrument_id)
            if not kvartal.empty and len(kvartal) >= 4:
                rad["inntjeningsstreak"] = beregn_inntjeningsstreak(kvartal)
                rad.update(beregn_vekst_metrikker(kvartal))

            # Hent kurser og beregn momentum
            kurser = hent_kurser(instrument_id)
            if not kurser.empty:
                rad["kursmomentum"] = beregn_kursmomentum(kurser)
                rad["snitt_daglig_omsetning"] = beregn_snitt_daglig_omsetning(kurser)

            alle_data.append(rad)

    # Steg 3: Bygg DataFrame og ekskluder selskaper med negativ EBITDA
    df = pd.DataFrame(alle_data)
    før_eksklusjon = len(df)
    if "ev_ebitda" in df.columns:
        df = df[df["ev_ebitda"].notna()]  # Ekskluder negativ EBITDA
    logger.info(f"Ekskluderte {før_eksklusjon - len(df)} selskaper med negativ EBITDA")

    # Ekskluder selskaper med markedsverdi under 1 mrd (lokal valuta)
    if "market_cap_mill" in df.columns:
        før = len(df)
        df = df[df["market_cap_mill"].notna() & (df["market_cap_mill"] > 1000)]
        logger.info(f"Ekskluderte {før - len(df)} selskaper med markedsverdi under 1 mrd")

    # Steg 4: Beregn scores
    df = beregn_samlet_score(df, FAKTORER)

    # Steg 5: Sorter og velg kolonner for output
    df = df.sort_values("samlet_score", ascending=False).reset_index(drop=True)

    output_kolonner = [
        "navn", "ticker", "sektor", "land",
        "samlet_score",
        "verdi_score", "kvalitet_score", "momentum_score", "vekst_score",
        "pe", "pb", "ev_ebitda", "ev_ebit", "roic", "roce",
        "kursmomentum", "inntjeningsstreak",
    ]

    # Inkluder bare kolonner som faktisk finnes i DataFrame
    output_kolonner = [k for k in output_kolonner if k in df.columns]
    df_output = df[output_kolonner]

    # Gi kolonnene lesbare navn
    kolonne_navn = {
        "navn": "Navn", "ticker": "Ticker", "sektor": "Sektor", "land": "Land",
        "samlet_score": "Samlet score", "verdi_score": "Verdi score",
        "kvalitet_score": "Kvalitet score", "momentum_score": "Momentum score",
        "vekst_score": "Vekst score",
    }
    df_output = df_output.rename(columns=kolonne_navn)

    # Steg 6: Vis per-marked-tabeller i terminalen med rich
    antall_vis = TEST_ANTALL if TEST_ANTALL is not None else 10
    visnings_kolonner = ["Navn", "Ticker", "Sektor", "Samlet score", "Verdi score", "Kvalitet score", "Momentum score", "Vekst score"]
    visnings_kolonner = [k for k in visnings_kolonner if k in df_output.columns]

    marked_titler = {"Norge": "TOPP NORGE", "Sverige": "TOPP SVERIGE", "Danmark": "TOPP DANMARK"}

    for land, tittel in marked_titler.items():
        df_land = df_output[df_output["Land"] == land]
        if df_land.empty:
            continue

        tabell = Table(title=tittel, show_lines=True)
        for kolonne in visnings_kolonner:
            tabell.add_column(kolonne, justify="right" if "score" in kolonne.lower() else "left")

        for _, rad in df_land[visnings_kolonner].head(antall_vis).iterrows():
            tabell.add_row(*[
                f"{v:.1f}" if isinstance(v, float) else str(v)
                for v in rad
            ])

        console.print(tabell)

    # Vis "oversett kvalitet" – negativ kursmomentum, sortert på kvalitet-score
    if "kursmomentum" in df.columns and "kvalitet_score" in df.columns:
        oversett = df[
            (df["kursmomentum"] < -0.10) &
            (df["roic"].notna()) & (df["roic"] > 0) &
            (df["roce"].notna()) & (df["roce"] > 0) &
            (df["market_cap_mill"].notna()) & (df["market_cap_mill"] > 1000) &
            (df["snitt_daglig_omsetning"].notna()) & (df["snitt_daglig_omsetning"] > 1_000_000)
        ].sort_values("kvalitet_score", ascending=False)
        if not oversett.empty:
            oversett_kolonner = ["navn", "ticker", "sektor", "land", "kvalitet_score", "kursmomentum", "roic", "roce"]
            oversett_kolonner = [k for k in oversett_kolonner if k in oversett.columns]
            oversett_display = oversett[oversett_kolonner].head(10)

            tabell2 = Table(title="OVERSETT KVALITET (negativ kursmomentum, høy ROIC/ROCE)", show_lines=True)
            kolonne_navn2 = {
                "navn": "Navn", "ticker": "Ticker", "sektor": "Sektor", "land": "Land",
                "kvalitet_score": "Kvalitet score", "kursmomentum": "Kursmomentum",
                "roic": "ROIC", "roce": "ROCE",
            }
            for k in oversett_kolonner:
                tabell2.add_column(kolonne_navn2.get(k, k), justify="right" if k not in ("navn", "ticker", "sektor", "land") else "left")

            for _, rad in oversett_display.iterrows():
                tabell2.add_row(*[
                    f"{v:.1%}" if k == "kursmomentum" and isinstance(v, float) else
                    f"{v:.1f}" if isinstance(v, float) else str(v)
                    for k, v in zip(oversett_kolonner, rad)
                ])

            console.print(tabell2)

    # Steg 7: Lagre til Excel med formatering
    score_kolonner = ["Samlet score", "Verdi score", "Kvalitet score", "Momentum score", "Vekst score"]
    excel_kolonner = ["Navn", "Ticker", "Sektor", "Land"] + [k for k in score_kolonner if k in df_output.columns]
    df_excel = df_output[excel_kolonner].copy()
    for k in score_kolonner:
        if k in df_excel.columns:
            df_excel[k] = df_excel[k].round(1)

    output_sti = os.path.join(OUTPUT_MAPPE, OUTPUT_FILNAVN.replace(".csv", ".xlsx"))

    from openpyxl import load_workbook
    from openpyxl.styles import Font
    if os.path.exists(output_sti):
        with pd.ExcelWriter(output_sti, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_excel.to_excel(writer, sheet_name="Resultater", index=False)
    else:
        with pd.ExcelWriter(output_sti, engine="openpyxl") as writer:
            df_excel.to_excel(writer, sheet_name="Resultater", index=False)

    wb = load_workbook(output_sti)
    # Fjern eventuelle auto-genererte "Sheet"-ark
    for navn in ["Sheet", "Sheet1"]:
        if navn in wb.sheetnames:
            del wb[navn]
    ws = wb["Resultater"]

    for row in ws.iter_rows():
        for celle in row:
            celle.font = Font(name="Arial", size=14)

    for celle in ws[1]:
        celle.font = Font(name="Arial", size=14, bold=True)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for celle in row:
            celle.font = Font(name="Arial", size=14, bold=True)

    for kolonne in ws.columns:
        bredde = max(len(str(celle.value)) if celle.value is not None else 0 for celle in kolonne)
        ws.column_dimensions[kolonne[0].column_letter].width = bredde * 1.4 + 4

    wb.save(output_sti)
    logger.info(f"Resultat lagret til {output_sti}")

    # Legg til kjøring i logg-arket
    legg_til_logg(df_output, FAKTORER, MARKEDER, output_sti)
    logger.info("Logg oppdatert")

    logger.info(f"=== Screening fullført – {len(df_output)} selskaper rangert ===")

    return df_output


if __name__ == "__main__":
    kjør_screening()
