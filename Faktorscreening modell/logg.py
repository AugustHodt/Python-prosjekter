
# Håndterer kjørelogg i Excel-filen.
# Legger til en ny fargekodet rad i "Logg"-arket for hver kjøring.
# Samme dato får samme farge – ny dato får neste farge i rekken.

from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

# Pastellfarger som roterer for hver ny dato
FARGER = [
    "D9E8FB",  # Lyseblå
    "D5F0D5",  # Lysegrønn
    "FFF2CC",  # Lysgul
    "FCE4EC",  # Lyserosa
    "FFE0CC",  # Lysoransje
    "EDE7F6",  # Lilla
    "E0F7FA",  # Turkis
    "F3E5F5",  # Lavendel
]

LOGG_ARK = "Logg"
LOGG_HEADER = (
    ["Dato", "Tid", "Aktive faktorer", "Marked"] +
    [f"#{i}" for i in range(1, 11)]
)


def legg_til_logg(df_output: pd.DataFrame, faktorer: dict, markeder: dict, output_sti: str):
    """
    Legger til en ny rad i Logg-arket i Excel-filen.
    Farger raden basert på datoen – ny dato = ny farge.
    """
    if not os.path.exists(output_sti):
        return

    wb = load_workbook(output_sti)

    # Opprett Logg-ark hvis det ikke finnes
    if LOGG_ARK not in wb.sheetnames:
        ws_logg = wb.create_sheet(LOGG_ARK)
        _skriv_header(ws_logg)
    else:
        ws_logg = wb[LOGG_ARK]
        # Legg til header hvis arket er tomt
        if ws_logg.max_row == 0 or ws_logg.cell(1, 1).value != "Dato":
            _skriv_header(ws_logg)

    # Bestem farge for dagens dato
    farge = _hent_farge_for_dato(ws_logg)

    # Bygg radverdier
    dato_nå = datetime.now()
    aktive_faktorer = [k.capitalize() for k, v in faktorer.items() if v]

    navn_kol = "Navn" if "Navn" in df_output.columns else "navn"
    land_kol = "Land" if "Land" in df_output.columns else "land"
    score_kol = "Samlet score" if "Samlet score" in df_output.columns else "samlet_score"

    def topp10_for_land(land_navn):
        df_land = df_output[df_output[land_kol] == land_navn]
        resultat = []
        for _, rad in df_land.head(10).iterrows():
            navn = rad.get(navn_kol, "–")
            score = rad.get(score_kol, "–")
            score_str = f"{score:.1f}" if isinstance(score, float) else str(score)
            resultat.append(f"{navn} ({score_str})")
        while len(resultat) < 10:
            resultat.append("–")
        return resultat

    fyll = PatternFill(start_color=farge, end_color=farge, fill_type="solid")
    dato_str = dato_nå.strftime("%Y-%m-%d")
    tid_str = dato_nå.strftime("%H:%M")
    faktorer_str = ", ".join(aktive_faktorer)

    # Én rad per aktivt marked
    markeder_rader = [
        ("Norge",   topp10_for_land("Norge")),
        ("Sverige", topp10_for_land("Sverige")),
        ("Danmark", topp10_for_land("Danmark")),
    ]

    for marked_navn, topp10 in markeder_rader:
        # Hopp over markeder som ikke er aktive (alle –)
        if all(v == "–" for v in topp10):
            continue

        rad_verdier = [dato_str, tid_str, faktorer_str, marked_navn] + topp10
        ny_rad = ws_logg.max_row + 1

        for col_idx, verdi in enumerate(rad_verdier, start=1):
            celle = ws_logg.cell(row=ny_rad, column=col_idx, value=verdi)
            celle.fill = fyll
            celle.font = Font(name="Arial", size=12)

    # Auto-juster kolonnebredde
    for col_idx in range(1, len(rad_verdier) + 1):
        maks_bredde = max(
            len(str(ws_logg.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws_logg.max_row + 1)
        )
        ws_logg.column_dimensions[get_column_letter(col_idx)].width = maks_bredde * 1.3 + 4

    wb.save(output_sti)


def _skriv_header(ws):
    """Skriver headerrad med fet skrift."""
    for col_idx, tittel in enumerate(LOGG_HEADER, start=1):
        celle = ws.cell(row=1, column=col_idx, value=tittel)
        celle.font = Font(name="Arial", size=12, bold=True)


def _hent_farge_for_dato(ws) -> str:
    """
    Finner riktig farge for dagens dato.
    Hvis datoen allerede finnes i loggen, bruk samme farge.
    Hvis det er en ny dato, bruk neste farge i rekken.
    """
    dato_i_dag = datetime.now().strftime("%Y-%m-%d")

    # Samle alle datoer som allerede er i loggen (hopp over header)
    datoer_sett = []
    for rad in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        dato = rad[0]
        if dato and dato not in datoer_sett:
            datoer_sett.append(dato)

    # Hvis dagens dato allerede er i loggen, bruk samme farge
    if dato_i_dag in datoer_sett:
        indeks = datoer_sett.index(dato_i_dag)
        return FARGER[indeks % len(FARGER)]

    # Ny dato – bruk neste farge
    return FARGER[len(datoer_sett) % len(FARGER)]
